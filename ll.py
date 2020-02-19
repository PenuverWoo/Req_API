import openpyxl
from openpyxl.workbook import Workbook
import datetime
from openpyxl import load_workbook

wb = load_workbook('ht.xlsx')
ws = wb['T']
ws['A1'] = 'as'
rows = [
 ['Number', 'data1', 'data2'],
 [2, 40, 30],
 [3, 40, 25],
 [4, 50, 30],
 [5, 30, 10],
 [6, 25, 5],
 [7, 50, 10],
]
print(delta=datetime.timedelta(days=3))
# print(list(zip(*rows)))
# ws.append(list(zip(*rows)))
# wb.save("ht.xlsx")
# ws.append([1,2,3])
# name_list = wb.get_sheet_names()
# print(name_list)
# outwb = Workbook()
# wo = outwb.active
#
# careerSheet = outwb.create_sheet('career',0)
# # careerSheet['A1']= datetime.datetime.now()
# careerSheet.cell(row=1,column=1).value = 'A'
# careerSheet.cell(row=1,column=2).value = 'B'
# careerSheet.cell(row=1,column=7).value = 'c'
#
# careerSheet.cell(row=2,column=2).value = 20
# careerSheet.cell(row=2,column=5).value= 30
# careerSheet.append([1,2,3])
# careerSheet.append(['This is A1', 'This is B1', 'This is C1'])
# careerSheet.append({'A' : 'This is A1', 'C' : 'This is C1'})
# careerSheet.append({1 : 'This is A1', 3 : 'This is C1'})

# outwb.save("sample.xlsx")